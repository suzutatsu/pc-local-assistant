import os
import asyncio
import logging
from dotenv import load_dotenv
from browser_use import Agent, Controller, Browser
from browser_use.llm.google import ChatGoogle
from google.auth import load_credentials_from_file
import openpyxl


# 環境変数の読み込み
load_dotenv()

# ナビゲーションのタイムアウトを延長するためのモンキーパッチ（デフォルトの4秒は短すぎるため）
# これは browser_use.browser.session.BrowserSession (Browser としてエイリアス) の内部メソッドをオーバーライドします
original_navigate_and_wait = Browser._navigate_and_wait

async def patched_navigate_and_wait(self, url, target_id, timeout=None):
    if timeout is None:
        # 遅いページのためにデフォルトの4秒のタイムアウトを20秒に延長
        timeout = 20.0
    return await original_navigate_and_wait(self, url, target_id, timeout)

Browser._navigate_and_wait = patched_navigate_and_wait

async def main():
    # Gemini Flashモデルの設定
    # ユーザー指定のモデル、もしくは最新のFlashモデル（gemini-3-flash-preview）を使用
    model_name = os.getenv("GEMINI_MODEL_NAME", "gemini-3-flash-preview")
    project_id = os.getenv("GOOGLE_CLOUD_PROJECT")
    location = os.getenv("GOOGLE_CLOUD_REGION", "asia-northeast1")
    credentials_path = os.getenv("GOOGLE_APPLICATION_CREDENTIALS")
    
    print(f"Using Vertex AI Model: {model_name}")
    print(f"Project: {project_id}, Region: {location}")
    print(f"Credentials Path: {credentials_path}")

    # パスが指定されている場合、認証情報を読み込む
    credentials = None
    if credentials_path and os.path.exists(credentials_path):
        credentials, _ = load_credentials_from_file(
            credentials_path,
            scopes=['https://www.googleapis.com/auth/cloud-platform']
        )
    
    # ChatGoogleを使用 (browser-useのネイティブコンポーネント)
    llm = ChatGoogle(
        model=model_name,
        vertexai=True,
        project=project_id,
        location=location,
        credentials=credentials,
        temperature=0
    )

    # タスク設定の読み込み
    import yaml
    import sys

    current_dir = os.getcwd()
    tasks_file = os.path.join(current_dir, "tasks.yaml")
    if not os.path.exists(tasks_file):
        print(f"エラー: 設定ファイル {tasks_file} が見つかりません。")
        return

    with open(tasks_file, 'r', encoding='utf-8') as f:
        try:
            config = yaml.safe_load(f)
            tasks = config.get('tasks', [])
        except yaml.YAMLError as exc:
            print(f"YAMLファイルの読み込みエラー: {exc}")
            return

    if not tasks:
        print("実行可能なタスクが定義されていません。")
        return

    selected_task = None
    
    # CLI引数を確認
    if len(sys.argv) > 1:
        arg = sys.argv[1]
        try:
            # 1始まりのインデックスを試行
            idx = int(arg) - 1
            if 0 <= idx < len(tasks):
                selected_task = tasks[idx]
        except ValueError:
            pass
            
    if selected_task:
        print(f"CLI引数でタスクが選択されました: {sys.argv[1]}")
    else:
        print("\n--- 実行可能タスク一覧 ---")
        for i, t in enumerate(tasks):
            print(f"{i + 1}. {t.get('name')} ({t.get('id')})")
            print(f"   説明: {t.get('description')}")
        print("-------------------------")

        # 対話モード
        while True:
            choice = input("実行したいタスクの番号を入力してください (qで終了): ")
            if choice.lower() == 'q':
                print("終了します。")
                return
            
            try:
                idx = int(choice) - 1
                if 0 <= idx < len(tasks):
                    selected_task = tasks[idx]
                    break
                else:
                    print("無効な番号です。もう一度入力してください。")
            except ValueError:
                print("番号を入力してください。")

    print(f"\n選択されたタスク: {selected_task.get('name')}")
    # ユーザーが「ログイン完了」等を伝えやすいように変数名をそのまま使用
    task_description = selected_task.get('prompt', '')

    # CLI引数からコンテキストを追加 (2つ目以降の引数がある場合)
    if len(sys.argv) > 2:
        extra_args = sys.argv[2:]
        context_info = " ".join(extra_args)
        print(f"追加コンテキスト: {context_info}")
        task_description += f"\n\n[ユーザー提供コンテキスト]: {context_info}\n(ここに情報が含まれている場合は、ask_userを使わずにそれを直接使用してください。)"

    if not task_description:
        print("タスクの説明（prompt）が空です。終了します。")
        return

    # ブラウザの仕様設定
    # use_browser が明示的に false の場合のみヘッドレスモードを有効化（画面を出さない）
    # デフォルトは True (画面を出す)
    use_browser = selected_task.get('use_browser', True)
    headless_mode = not use_browser

    # 永続的なブラウザプロファイルの設定
    # プロジェクトディレクトリ内に 'browser_profile' ディレクトリを作成して使用します
    profile_path = os.path.join(current_dir, "browser_profile")
    os.makedirs(profile_path, exist_ok=True)
    
    # ブラウザの初期化 (0.11.3以降のAPI)
    # Browser (BrowserSession) に直接設定を渡します
    browser = Browser(
        headless=headless_mode, 
        user_data_dir=profile_path,
        enable_default_extensions=False, # 組織ポリシー等で拡張機能インストールが制限される場合に備え無効化
        # chrome_instance_path=None, # デフォルトのChromeを使用
        # other args...
    )




    # Controller パターンを使用してカスタムアクション（ツール）を定義
    controller = Controller()
    
    @controller.action("ask_user")
    def ask_user_action(question: str):
        """
        ユーザーに質問をし、その回答を返します。
        MFAコード、認証コード、OTP、またはその他の情報を入力する必要がある場合にこのツールを使用してください。
        また、ログインなどの手動操作をユーザーに依頼する場合にも使用してください。
        """
        print(f"\n\n[Agentからの質問]: {question}")
        return input("回答を入力してください (入力後Enter): ")

    @controller.action("create_excel_from_template")
    def create_excel_from_template(template_path: str, output_path: str, updates_json: str):
        """
        テンプレートから新しいExcelファイルを作成し、特定のセルを更新します。
        
        Args:
            template_path: テンプレートExcelファイル(.xlsx)への絶対パス
            output_path: 新しいExcelファイルを保存する場所への絶対パス
            updates_json: 'SheetName!CellAddress' (例: 'Sheet1!A1') をキーとし、
                          そのセルの新しい内容を値とする辞書を表すJSON文字列。
                          例: '{"Sheet1!A1": "Hello", "Sheet1!B2": "World"}'
                     
        Returns:
            成功またはエラーメッセージを示す文字列。
        """
        try:
            import json
            
            try:
                updates = json.loads(updates_json)
            except json.JSONDecodeError as e:
                return f"Error: Failed to parse updates_json. Make sure it is a valid JSON string. Error: {e}"
            # Resolve relative paths
            if not os.path.isabs(template_path):
                template_path = os.path.abspath(os.path.join(os.getcwd(), template_path))
            if not os.path.isabs(output_path):
                output_path = os.path.abspath(os.path.join(os.getcwd(), output_path))

            if not os.path.exists(template_path):
                return f"Error: Template file not found at {template_path}"
            
            wb = openpyxl.load_workbook(template_path)
            
            for location, value in updates.items():
                if '!' in location:
                    sheet_name, cell_address = location.split('!', 1)
                    if sheet_name in wb.sheetnames:
                        ws = wb[sheet_name]
                    else:
                        return f"Error: Sheet '{sheet_name}' not found in template."
                else:
                    # Default to active sheet if no sheet specified
                    ws = wb.active
                    cell_address = location
                
                ws[cell_address] = value
                
            # Ensure directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            wb.save(output_path)
            return f"Successfully created Excel file at {output_path} with updated values."
            
        except Exception as e:
            return f"Error processing Excel file: {str(e)}"

    # エージェントの作成
    agent = Agent(
        task=task_description,
        llm=llm,
        browser=browser,
        controller=controller
    )

    # エージェントの実行
    print("エージェントを実行中...")
    history = await agent.run()
    
    # 結果の表示
    print("\n--- 実行結果 ---")
    print(history.final_result())

    # ブラウザを閉じる
    # await browser.close() # browser-use 0.11.3 handles cleanup automatically

if __name__ == "__main__":
    asyncio.run(main())
